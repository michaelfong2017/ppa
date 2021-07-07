#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Dec 29 23:01:35 2020

@author: michael
"""
# %%
# import for database
import psycopg2

# import for filter
import re

# import for tokenize_and_store
import os

# import for logging
from util.logger import create_logger
import datetime

# Gensim
import gensim  # conda install -c conda-forge gensim
import gensim.corpora as corpora
from gensim.utils import simple_preprocess
from gensim.models import CoherenceModel

import pandas as pd
import numpy as np

from pprint import pprint

# Parallel computing
from util.Parallelizer import make_parallel

# Save result file
import openpyxl

# %%
def setup():
    FILE_DIR = os.path.dirname(os.path.abspath(__file__))

    global logger
    logger = create_logger()

    #
    class TopicModel:
        NUMBER_OF_RECORDS = 0
        OFFSET = 0
        user = None
        passwd = None

        def __init__(self, number_of_records, offset):
            self.NUMBER_OF_RECORDS = number_of_records
            self.OFFSET = offset

        @staticmethod
        def reset_credentials():
            TopicModel.user = None
            TopicModel.passwd = None

        def get_data_words_list(self):
            """
            Retrieve data from PostgreSQL
            """
            if TopicModel.user is None:
                TopicModel.user = input("Enter PostgreSQL username: ")
            if TopicModel.passwd is None:
                TopicModel.passwd = input("Enter PostgreSQL user password: ")

            conn = psycopg2.connect(
                f"dbname=budgetq user={TopicModel.user} host=dev.clo3yq4mhvjy.ap-east-1.rds.amazonaws.com password={TopicModel.passwd}"
            )

            if logger is not None:
                logger.info(f"NUMBER_OF_RECORDS: {self.NUMBER_OF_RECORDS}")
                logger.info(f"OFFSET: {self.OFFSET}")

            start_time = datetime.datetime.now()

            try:
                with conn.cursor() as cursor:
                    cursor.execute(
                        f"select tokenized_question, tokenized_answer from question where id > {self.OFFSET} ORDER BY id LIMIT {self.NUMBER_OF_RECORDS}"
                    )
                    records = cursor.fetchall()

                    if logger is not None:
                        logger.info(
                            f"Time elapsed for fetching records: {datetime.datetime.now() - start_time}"
                        )
                    start_time = datetime.datetime.now()

                    cursor.close()

            finally:
                if conn:
                    conn.close()

            def format_records(q_a):
                return [col.rstrip(";").split(";") for col in q_a]

            data_words_list = list(map(format_records, zip(*records)))

            if logger is not None:
                logger.info(
                    f"Time elapsed for formatting records {self.OFFSET + 1} to {self.OFFSET + self.NUMBER_OF_RECORDS}: {datetime.datetime.now() - start_time}"
                )

            return data_words_list

    #
    # TopicModel.reset_credentials()

    #
    topic_model = TopicModel(100000, 0)
    global data_words_list
    data_words_list = topic_model.get_data_words_list()
    logger.info(np.shape(data_words_list))


def load_params():
    wb_obj = openpyxl.load_workbook("train_params.xlsx")
    ws = wb_obj.active
    return [
        tuple(cell.value for cell in row) for row in ws.iter_rows(max_row=ws.max_row)
    ]


# Prevent rebuilding dictionary, which is very time-consuming.
def make_corpora(data_words, filter_no_above):
    # Build the bigram and trigram models
    bigram = gensim.models.Phrases(
        data_words, min_count=5, threshold=100
    )  # higher threshold fewer phrases.
    trigram = gensim.models.Phrases(bigram[data_words], threshold=100)

    # Faster way to get a sentence clubbed as a trigram/bigram
    bigram_mod = gensim.models.phrases.Phraser(bigram)
    trigram_mod = gensim.models.phrases.Phraser(trigram)

    # See trigram example
    # print(bigram_mod[data_words[0]])

    def lemmatized_to_corpus(data_lemmatized):
        # Create Dictionary
        id2word = corpora.Dictionary(data_lemmatized)
        """
        Use no_above to filter stopwords, which are very frequent
        """
        id2word.filter_extremes(no_above=filter_no_above)

        # Create Corpus
        texts = data_lemmatized

        # Term Document Frequency
        corpus = [id2word.doc2bow(text) for text in texts]

        # View
        logger.info(corpus[:1])

        #
        logger.info([[(id2word[id], freq) for id, freq in cp] for cp in corpus[:2]])

        return corpus, id2word

    data_lemmatized = data_words
    bigram_data_lemmatized = [bigram_mod[doc] for doc in data_words]
    trigram_data_lemmatized = [trigram_mod[doc] for doc in data_words]

    corpus, id2word = lemmatized_to_corpus(data_lemmatized)
    bigram_corpus, bigram_id2word = lemmatized_to_corpus(bigram_data_lemmatized)
    trigram_corpus, trigram_id2word = lemmatized_to_corpus(trigram_data_lemmatized)

    return (
        data_lemmatized,
        corpus,
        id2word,
        bigram_data_lemmatized,
        bigram_corpus,
        bigram_id2word,
        trigram_data_lemmatized,
        trigram_corpus,
        trigram_id2word,
    )


#
def train(
    data_lemmatized,
    corpus,
    id2word,
    use_bigram,
    use_trigram,
    filter_no_above,
    num_topics,
    alpha,
    eta,
    random_state,
    passes,
    save_filename,
    topn=20,
):
    #
    # Build LDA model
    start_time = datetime.datetime.now()

    lda_model = gensim.models.ldamodel.LdaModel(
        corpus=corpus,
        id2word=id2word,
        num_topics=num_topics,
        random_state=random_state,
        update_every=1,
        chunksize=100,
        passes=passes,
        # alpha=np.full(num_topics, 0.01),
        # eta=np.full(num_topics, 0.01),
        alpha=alpha,
        eta=eta,
        per_word_topics=True,
    )

    time_elapsed = datetime.datetime.now() - start_time
    logger.info(f"Time elapsed for training LDA model: {time_elapsed}")
    #
    # Print the Keyword in the 20 topics
    for i in range(20):
        pprint(lda_model.print_topic(i))
    doc_lda = lda_model[corpus]

    #
    # Compute Perplexity
    perplexity_lda = lda_model.log_perplexity(corpus)
    print(
        "\nPerplexity: ", perplexity_lda
    )  # a measure of how good the model is. lower the better.

    # Compute Coherence Score
    coherence_model_lda = CoherenceModel(
        model=lda_model, texts=data_lemmatized, dictionary=id2word, coherence="c_v"
    )
    coherence_lda = coherence_model_lda.get_coherence()
    print("\nCoherence Score: ", coherence_lda)

    #
    wb = openpyxl.Workbook()
    ws_write = wb.active
    for j in range(num_topics):
        all_words = lda_model.print_topic(j, topn=topn).split(" + ")
        for i in range(topn):
            cell = ws_write.cell(row=i + 1, column=j + 1)
            cell.value = all_words[i]

    ws_write.cell(row=topn + 2, column=1).value = "Perplexity: " + str(
        perplexity_lda
    )
    ws_write.cell(row=topn + 3, column=1).value = "Coherence Score: " + str(
        coherence_lda
    )

    ws_write.cell(
        row=topn + 5, column=1
    ).value = "Time elapsed for training LDA model: " + str(time_elapsed)

    ws_write.cell(row=topn + 7, column=1).value = "use_bigram: " + str(use_bigram)
    ws_write.cell(row=topn + 8, column=1).value = "use_trigram: " + str(
        use_trigram
    )
    ws_write.cell(row=topn + 9, column=1).value = "filter_no_above: " + str(
        filter_no_above
    )
    ws_write.cell(row=topn + 10, column=1).value = "num_topics: " + str(
        num_topics
    )
    ws_write.cell(row=topn + 11, column=1).value = (
        "alpha: "
        + np.array2string(alpha, formatter={"float_kind": lambda x: "%.4f" % x})
        if type(alpha) is np.ndarray
        else str(alpha)
    )
    ws_write.cell(row=topn + 12, column=1).value = (
        "eta: " + np.array2string(eta, formatter={"float_kind": lambda x: "%.4f" % x})
        if type(eta) is np.ndarray
        else str(eta)
    )
    ws_write.cell(row=topn + 13, column=1).value = "random_state: " + str(
        random_state
    )
    ws_write.cell(row=topn + 14, column=1).value = "passes: " + str(passes)
    ws_write.cell(row=topn + 15, column=1).value = (
        "save_filename: " + save_filename
    )

    wb.save(filename=save_filename)

    return perplexity_lda, coherence_lda


# %%
def train_with_params(
    data_words_type,
    num_topics,
    use_bigram,
    use_trigram,
    filter_no_above,
    alpha_entry,
    eta_entry,
    random_state,
    passes,
    topn,
    data_lemmatized,
    corpus,
    id2word,
    bigram_data_lemmatized,
    bigram_corpus,
    bigram_id2word,
    trigram_data_lemmatized,
    trigram_corpus,
    trigram_id2word,
):
    # [50,100,200,300,400,500,600,1000]
    alpha = "auto" if alpha_entry == "auto" else np.full(num_topics, alpha_entry)
    eta = "auto" if eta_entry == "auto" else None

    folder_path = f"output/n_{num_topics}/bi_{use_bigram}/tri_{use_trigram}/"
    os.makedirs(folder_path, exist_ok=True)
    save_filename = (
        f"{folder_path}fna_{filter_no_above}_alpha_{alpha_entry}_eta_{eta_entry}.xlsx"
    )

    if use_bigram:
        lemma = bigram_data_lemmatized
        cor = bigram_corpus
        id2 = bigram_id2word
    elif use_trigram:
        lemma = trigram_data_lemmatized
        cor = trigram_corpus
        id2 = trigram_id2word
    else:
        lemma = data_lemmatized
        cor = corpus
        id2 = id2word

    perplexity, coherence_score = train(
        data_lemmatized=lemma,
        corpus=cor,
        id2word=id2,
        num_topics=num_topics,
        use_bigram=use_bigram,
        use_trigram=use_trigram,
        filter_no_above=filter_no_above,
        alpha=alpha,
        eta=eta,
        random_state=random_state,
        passes=passes,
        save_filename=save_filename,
        topn=topn,
    )

    # Save results to db
    conn = psycopg2.connect(
        f"dbname=budgetq user={db_user} host=dev.clo3yq4mhvjy.ap-east-1.rds.amazonaws.com password={db_passwd}"
    )

    try:
        with conn.cursor() as cursor:
            try:
                cursor.execute(
                    f"""INSERT INTO training ("data_words_type", "num_topics", "use_bigram", "use_trigram", "filter_no_above", "alpha_entry", "eta_entry", "random_state", "passes", "topn", "perplexity", "coherence_score") VALUES ('{data_words_type}', '{num_topics}', '{use_bigram}', '{use_trigram}', '{filter_no_above}', '{alpha_entry}', '{eta_entry}', '{random_state}', '{passes}', '{topn}', '{perplexity}', '{coherence_score}')
                    ON CONFLICT ("data_words_type", "num_topics", "use_bigram", "use_trigram", "filter_no_above", "alpha_entry", "eta_entry", "random_state", "passes", "topn") DO UPDATE 
                    SET "data_words_type" = '{data_words_type}', 
                    "num_topics" = '{num_topics}',
                    "use_bigram" = '{use_bigram}',
                    "use_trigram" = '{use_trigram}',
                    "filter_no_above" = '{filter_no_above}',
                    "alpha_entry" = '{alpha_entry}',
                    "eta_entry" = '{eta_entry}',
                    "random_state" = '{random_state}',
                    "passes" = '{passes}',
                    "topn" = '{topn}',
                    "perplexity" = '{perplexity}',
                    "coherence_score" = '{coherence_score}';"""
                )
                conn.commit()
            except Exception as e:
                logger.error(f"{e}")
                conn.rollback()
                return f"{e}"
    finally:
        if conn:
            conn.close()


# %%
if __name__ == "__main__":
    setup()

# %%
if __name__ == "__main__":
    COMPLETED_N = 41

    params = load_params()
    logger.info(params[0])

    global db_user, db_passwd
    db_user = input("Enter PostgreSQL username: ")
    db_passwd = input("Enter PostgreSQL user password: ")

    # Group each question and each answer together to improve the score
    # data_words = list(map(lambda t: t[0] + t[1], zip(data_words_list[0], data_words_list[1])))
    # data_words = data_words_list[0] + data_words_list[1]
    data_words = list(
        map(lambda t: t[0] + t[1], zip(data_words_list[0], data_words_list[1]))
    )

    # mp_train_with_params = make_parallel(
    #     type="multiprocessing",
    #     has_different_tasks=True,
    #     has_multiple_arguments=True,
    #     max_workers=8,
    # )(train_with_params)
    if len(params) > 1:
        current_fna = params[1 + COMPLETED_N][4]
        (
            data_lemmatized,
            corpus,
            id2word,
            bigram_data_lemmatized,
            bigram_corpus,
            bigram_id2word,
            trigram_data_lemmatized,
            trigram_corpus,
            trigram_id2word,
        ) = make_corpora(data_words, current_fna)

# %%
if __name__ == "__main__":
    for i in range(len(params)):
        if i > 0 + COMPLETED_N:
            if not params[i][4] == current_fna:
                current_fna = params[i][4]
                (
                    data_lemmatized,
                    corpus,
                    id2word,
                    bigram_data_lemmatized,
                    bigram_corpus,
                    bigram_id2word,
                    trigram_data_lemmatized,
                    trigram_corpus,
                    trigram_id2word,
                ) = make_corpora(data_words, current_fna)

            train_with_params(
                *params[i],
                data_lemmatized,
                corpus,
                id2word,
                bigram_data_lemmatized,
                bigram_corpus,
                bigram_id2word,
                trigram_data_lemmatized,
                trigram_corpus,
                trigram_id2word,
            )
# %%
